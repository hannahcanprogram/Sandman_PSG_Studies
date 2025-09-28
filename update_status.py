# update_status.py —— openpyxl
import sys, csv, time, os
from openpyxl import load_workbook

LOG_CSV = r"E:\Machine_01\Sandman\Logs\sandman_batch_log_group4A.csv"

def to_str(x):
    return "" if x is None else str(x).strip()

def find_col(ws, header):
    tgt = header.lower()
    for c in range(1, ws.max_column + 1):
        if to_str(ws.cell(1, c).value).lower() == tgt:
            return c
    raise ValueError(f"Column '{header}' not found.")

def open_wb_retry(path, tries=10, delay=0.4):
    for i in range(tries):
        try:
            return load_workbook(path, read_only=False, data_only=True)
        except PermissionError:
            time.sleep(delay)
        except Exception:
            if i == tries - 1:
                raise
            time.sleep(delay)

def save_replace_retry(wb, path, tries=10, delay=0.4):
    tmp = path + ".tmp"
    for i in range(tries):
        try:
            wb.save(tmp)
            os.replace(tmp, path)
            return
        except PermissionError:
            time.sleep(delay)
        except Exception:
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except:
                pass
            if i == tries - 1:
                raise
            time.sleep(delay)

def append_log(row, name, path, status, extra):
    try:
        os.makedirs(os.path.dirname(LOG_CSV), exist_ok=True)
        newfile = (not os.path.exists(LOG_CSV)) or os.path.getsize(LOG_CSV) == 0
        with open(LOG_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if newfile:
                w.writerow(["timestamp", "row", "name", "path", "status", "info"])
            w.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S"),
                row, name, path, status, extra
            ])
    except Exception:
        pass

def main(xlsx, sheet, row, status, extra):
    wb = open_wb_retry(xlsx)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet}' not found.")
        ws = wb[sheet]

        c_stat = find_col(ws, "Status")
        c_name = find_col(ws, "Name")
        c_path = find_col(ws, "Path")

        if row < 2 or row > ws.max_row:
            raise ValueError(f"Row {row} out of range (max={ws.max_row}).")

        if status.lower() == "success":
            ws.cell(row, c_stat).value = "1"
        else:
            ws.cell(row, c_stat).value = ("error:" + to_str(extra))[:255]

        name = to_str(ws.cell(row, c_name).value)
        path = to_str(ws.cell(row, c_path).value)

        save_replace_retry(wb, xlsx)
    finally:
        try:
            wb.close()
        except:
            pass

    append_log(row, name, path, status, to_str(extra))

if __name__ == "__main__":
    try:
        xlsx  = sys.argv[1]
        sheet = sys.argv[2]
        row   = int(sys.argv[3])
        status= sys.argv[4]            # success / error
        extra = sys.argv[5] if len(sys.argv) > 5 else ""
        main(xlsx, sheet, row, status, extra)
    except Exception as e:
        sys.stderr.write("ERROR: " + repr(e) + "\n")
        sys.stderr.flush()
        sys.exit(1)
