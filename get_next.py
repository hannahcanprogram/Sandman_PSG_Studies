# get_next.py  —— openpyxl
import sys, json, time, os
from openpyxl import load_workbook

def to_str(v):
    if v is None: return ""
    return str(v).strip()

def is_na(s):
    s = to_str(s).lower()
    return s in ("", "na", "n/a", "none", "nan")

def is_one(v):
    s = to_str(v).lower()
    if s == "": return False
    try:
        return float(s) == 1.0
    except Exception:
        return s in {"done", "complete"}

def should_skip(v):
    s = to_str(v).lower()
    return is_one(s) or s == "running"

def find_col(ws, header):
    tgt = header.lower()
    for c in range(1, ws.max_column + 1):
        if to_str(ws.cell(1, c).value).lower() == tgt:
            return c
    raise ValueError(f"Column '{header}' not found in row 1.")

def open_wb_retry(path, tries=10, delay=0.4):
    for i in range(tries):
        try:
            return load_workbook(path, read_only=False, data_only=True)
        except PermissionError:
            time.sleep(delay)
        except Exception:
            if i == tries - 1: raise
            time.sleep(delay)

def save_replace_retry(wb, path, tries=10, delay=0.4):
    tmp = path + ".tmp"
    for i in range(tries):
        try:
            wb.save(tmp)
            os.replace(tmp, path)
            return
        except PermissionError:
            time.sleep(delay)
        except Exception:
            try:
                if os.path.exists(tmp): os.remove(tmp)
            except: pass
            if i == tries - 1: raise
            time.sleep(delay)

def main(xlsx, sheet_name="Sheet1"):
    wb = open_wb_retry(xlsx)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found.")
        ws = wb[sheet_name]

        c_path = find_col(ws, "Path")
        c_name = find_col(ws, "Name")
        c_stat = find_col(ws, "Status")

        last = ws.max_row
        while last >= 2 and to_str(ws.cell(last, c_path).value) == "":
            last -= 1
        if last < 2:
            print("{}"); sys.stdout.flush(); return

        target_row = None
        for r in range(2, last + 1):
            if not should_skip(ws.cell(r, c_stat).value):
                target_row = r
                break

        if not target_row:
            print("{}"); sys.stdout.flush(); return

        p = ws.cell(target_row, c_path).value
        n = ws.cell(target_row, c_name).value

        ws.cell(target_row, c_stat).value = "running"
        save_replace_retry(wb, xlsx)

        if is_na(p):
            print(json.dumps({"path": "na"})); sys.stdout.flush(); return

        out = {"path": to_str(p), "name": to_str(n), "row": int(target_row)}
        print(json.dumps(out, ensure_ascii=False)); sys.stdout.flush()
    finally:
        try: wb.close()
        except: pass

if __name__ == "__main__":
    try:
        xlsx = sys.argv[1]
        sheet = sys.argv[2] if len(sys.argv) > 2 else "Sheet1"
        main(xlsx, sheet)
    except Exception as e:
        sys.stderr.write("ERROR: " + repr(e) + "\n")
        sys.stderr.flush()
        sys.exit(1)
